"""Excel 全功能展示文档 - openpyxl 能力一览"""
import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import CellIsRule, DataBarRule, ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ====== 配色：奶酪色 × 蒂芙尼蓝 ======
TIFFANY = "81D8D1"
DEEP_TF = "5CBDB4"
CREAM    = "FBF7D8"
LIGHT_CR = "FDF9EC"
MINT     = "A8E0D8"
DARK     = "2B2D42"
GRAY     = "7A7C8E"
WHITE    = "FFFFFF"

# ====== 样式预制 ======
header_font = Font(name="Microsoft YaHei", bold=True, color=WHITE, size=12)
header_fill = PatternFill(start_color=TIFFANY, end_color=TIFFANY, fill_type="solid")
sub_header_fill = PatternFill(start_color=DEEP_TF, end_color=DEEP_TF, fill_type="solid")
cream_fill = PatternFill(start_color=CREAM, end_color=CREAM, fill_type="solid")
light_fill = PatternFill(start_color=LIGHT_CR, end_color=LIGHT_CR, fill_type="solid")
mint_fill = PatternFill(start_color=MINT, end_color=MINT, fill_type="solid")

normal_font = Font(name="Microsoft YaHei", color=DARK, size=11)
gray_font = Font(name="Microsoft YaHei", color=GRAY, size=10)
title_font = Font(name="Microsoft YaHei", bold=True, color=DARK, size=16)
section_font = Font(name="Microsoft YaHei", bold=True, color=DEEP_TF, size=13)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="D0C8BC"),
    right=Side(style="thin", color="D0C8BC"),
    top=Side(style="thin", color="D0C8BC"),
    bottom=Side(style="thin", color="D0C8BC"),
)

thick_bottom = Border(bottom=Side(style="medium", color=TIFFANY))


def style_header_row(ws, row, cols):
    """设置表头行样式"""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border


def style_data_rows(ws, start_row, end_row, cols):
    """设置数据行交替色"""
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = normal_font
            cell.alignment = center_align
            cell.border = thin_border
            if (r - start_row) % 2 == 1:
                cell.fill = light_fill


def write_section_title(ws, row, text, cols=8):
    """写入章节标题（合并单元格）"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = section_font
    cell.alignment = Alignment(horizontal="left", vertical="center")
    # 底部装饰线
    cell.border = thick_bottom


# ========================================
# 创建工作簿
# ========================================
wb = Workbook()


# ========================================
# Sheet 1：文字与格式
# ========================================
ws1 = wb.active
ws1.title = "文字格式"
ws1.sheet_properties.tabColor = TIFFANY

# 列宽
widths = [16, 16, 16, 16, 16, 16, 16, 16]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# 标题
ws1.merge_cells("A1:H1")
c = ws1["A1"]
c.value = "openpyxl 全功能展示"
c.font = title_font
c.alignment = center_align

# 装饰线
ws1.merge_cells("A2:H2")
c = ws1["A2"]
c.fill = PatternFill(start_color=TIFFANY, end_color=TIFFANY, fill_type="solid")

# --- 字号对比 ---
row = 4
write_section_title(ws1, row, "字号对比")
row += 1

sizes = [
    ("20pt 大标题", 20, True),
    ("16pt 章节标题", 16, True),
    ("14pt 副标题", 14, True),
    ("12pt 正文", 12, False),
    ("10pt 小字", 10, False),
    ("9pt 注释", 9, False),
]

headers = ["效果", "字号", "加粗"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=row, column=i, value=h)
style_header_row(ws1, row, 3)
row += 1

for text, size, bold in sizes:
    ws1.cell(row=row, column=1, value=text).font = Font(
        name="Microsoft YaHei", size=size, bold=bold, color=DARK
    )
    ws1.cell(row=row, column=2, value=f"{size}pt").font = normal_font
    ws1.cell(row=row, column=3, value="是" if bold else "否").font = normal_font
    for c in range(1, 4):
        ws1.cell(row=row, column=c).alignment = center_align
        ws1.cell(row=row, column=c).border = thin_border
    row += 1

row += 1

# --- 颜色展示 ---
write_section_title(ws1, row, "颜色展示")
row += 1

colors = [
    ("蒂芙尼蓝", TIFFANY, "表头、主色"),
    ("深蒂蓝", DEEP_TF, "深色装饰"),
    ("奶酪色", CREAM, "页面背景"),
    ("浅奶黄", LIGHT_CR, "交替行色"),
    ("薄荷绿", MINT, "中间色"),
    ("深色文字", DARK, "正文"),
    ("灰色", GRAY, "次要信息"),
]

headers = ["色名", "色值", "用途", "预览"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=row, column=i, value=h)
style_header_row(ws1, row, 4)
row += 1

for name, hex_color, usage in colors:
    ws1.cell(row=row, column=1, value=name).font = normal_font
    ws1.cell(row=row, column=2, value=f"#{hex_color}").font = normal_font
    ws1.cell(row=row, column=3, value=usage).font = normal_font
    preview = ws1.cell(row=row, column=4)
    preview.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
    preview.value = "  "
    for c in range(1, 5):
        ws1.cell(row=row, column=c).alignment = center_align
        ws1.cell(row=row, column=c).border = thin_border
    row += 1

row += 1

# --- 对齐方式 ---
write_section_title(ws1, row, "对齐方式")
row += 1

aligns = [
    ("左对齐", "left"),
    ("居中对齐", "center"),
    ("右对齐", "right"),
    ("两端对齐（自动换行）", "justify"),
]

headers = ["效果", "对齐方式", "属性值"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=row, column=i, value=h)
style_header_row(ws1, row, 3)
row += 1

for text, align_type in aligns:
    ws1.cell(row=row, column=1, value=text).font = normal_font
    ws1.cell(row=row, column=2, value=align_type).font = normal_font
    ws1.cell(row=row, column=3, value=f'Alignment(horizontal="{align_type}")').font = gray_font
    for c in range(1, 4):
        ws1.cell(row=row, column=c).alignment = center_align
        ws1.cell(row=row, column=c).border = thin_border
    row += 1

row += 1

# --- 数字格式 ---
write_section_title(ws1, row, "数字格式")
row += 1

formats = [
    ("3.14159", "0.00", "保留两位小数"),
    ("12345.67", "#,##0.00", "千分位"),
    ("0.856", "0.0%", "百分比"),
    ("168", "¥#,##0", "人民币"),
    ("2026-06-14", "YYYY-MM-DD", "日期"),
    ("14:30:00", "HH:MM:SS", "时间"),
]

headers = ["原始值", "格式", "说明", "效果"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=row, column=i, value=h)
style_header_row(ws1, row, 4)
row += 1

for val, fmt, desc in formats:
    ws1.cell(row=row, column=1, value=val).font = normal_font
    ws1.cell(row=row, column=2, value=fmt).font = normal_font
    ws1.cell(row=row, column=3, value=desc).font = normal_font
    # 尝试转为数字
    try:
        num_val = float(val)
        cell = ws1.cell(row=row, column=4, value=num_val)
    except ValueError:
        cell = ws1.cell(row=row, column=4, value=val)
    cell.number_format = fmt
    cell.font = normal_font
    for c in range(1, 5):
        ws1.cell(row=row, column=c).alignment = center_align
        ws1.cell(row=row, column=c).border = thin_border
    row += 1


# ========================================
# Sheet 2：表格
# ========================================
ws2 = wb.create_sheet("表格")
ws2.sheet_properties.tabColor = DEEP_TF

for i in range(1, 9):
    ws2.column_dimensions[get_column_letter(i)].width = 16

# 标题
ws2.merge_cells("A1:H1")
ws2["A1"].value = "表格与数据处理"
ws2["A1"].font = title_font
ws2["A1"].alignment = center_align

ws2.merge_cells("A2:H2")
ws2["A2"].fill = PatternFill(start_color=TIFFANY, end_color=TIFFANY, fill_type="solid")

# --- 基础表格 ---
row = 4
write_section_title(ws2, row, "销售数据表")
row += 1

sales_data = [
    ["月份", "产品A", "产品B", "产品C", "合计", "增长率"],
    ["1月", 12000, 8500, 6200, None, None],
    ["2月", 15000, 9200, 7100, None, None],
    ["3月", 13500, 11000, 8300, None, None],
    ["4月", 18000, 10500, 9500, None, None],
    ["5月", 21000, 12800, 11200, None, None],
    ["6月", 19500, 14200, 10800, None, None],
]

headers = sales_data[0]
cols = len(headers)
for i, h in enumerate(headers, 1):
    ws2.cell(row=row, column=i, value=h)
style_header_row(ws2, row, cols)
row += 1

for r_idx, data_row in enumerate(sales_data[1:]):
    for c_idx, val in enumerate(data_row):
        cell = ws2.cell(row=row, column=c_idx + 1, value=val)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
        if c_idx >= 1 and c_idx <= 3:
            cell.number_format = "#,##0"
        if r_idx % 2 == 1:
            cell.fill = light_fill
    # 合计公式（E列 = B+C+D）
    ws2.cell(row=row, column=5).value = f"=SUM(B{row}:D{row})"
    ws2.cell(row=row, column=5).number_format = "#,##0"
    ws2.cell(row=row, column=5).font = Font(name="Microsoft YaHei", bold=True, color=DARK, size=11)
    ws2.cell(row=row, column=5).alignment = center_align
    ws2.cell(row=row, column=5).border = thin_border
    # 增长率公式（F列）
    if r_idx > 0:
        ws2.cell(row=row, column=6).value = f"=(E{row}-E{row-1})/E{row-1}"
        ws2.cell(row=row, column=6).number_format = "0.0%"
    ws2.cell(row=row, column=6).font = normal_font
    ws2.cell(row=row, column=6).alignment = center_align
    ws2.cell(row=row, column=6).border = thin_border
    row += 1

row += 1

# --- 条件格式 ---
write_section_title(ws2, row, "条件格式（增长率着色）")
row += 1
ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
ws2.cell(row=row, column=1, value="增长率 > 10% 标绿，< 0% 标红").font = gray_font
row += 1

# 条件格式区域
data_end_row = row - 3  # 上面数据的结束行
ws2.conditional_formatting.add(
    f"F7:F{data_end_row}",
    CellIsRule(operator="greaterThan", formula=["0.1"],
               fill=PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
               font=Font(color="2E7D32")),
)
ws2.conditional_formatting.add(
    f"F7:F{data_end_row}",
    CellIsRule(operator="lessThan", formula=["0"],
               fill=PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
               font=Font(color="C62828")),
)

row += 2

# --- 数据条 ---
write_section_title(ws2, row, "数据条（产品A销售额）")
row += 1

ws2.conditional_formatting.add(
    f"B7:B{data_end_row}",
    DataBarRule(start_type="min", end_type="max", color=TIFFANY),
)

row += 2

# --- 色阶 ---
write_section_title(ws2, row, "色阶（产品B销售额）")
row += 1

ws2.conditional_formatting.add(
    f"C7:C{data_end_row}",
    ColorScaleRule(
        start_type="min", start_color="FDF9EC",
        mid_type="percentile", mid_value=50, mid_color=MINT,
        end_type="max", end_color=TIFFANY,
    ),
)

row += 2

# --- 数据验证 ---
write_section_title(ws2, row, "数据验证（下拉选择）")
row += 1

dv = DataValidation(type="list", formula1='"待处理,进行中,已完成"', allow_blank=True)
dv.error = "请从下拉列表中选择"
dv.errorTitle = "输入错误"
dv.prompt = "选择状态"
dv.promptTitle = "状态"

ws2.add_data_validation(dv)

status_data = [
    ["任务", "负责人", "状态"],
    ["需求评审", "张三", "已完成"],
    ["UI设计", "李四", "进行中"],
    ["后端开发", "王五", "待处理"],
    ["测试验收", "赵六", "待处理"],
]

for i, h in enumerate(status_data[0], 1):
    ws2.cell(row=row, column=i, value=h)
style_header_row(ws2, row, 3)
row += 1

for task, person, status in status_data[1:]:
    ws2.cell(row=row, column=1, value=task).font = normal_font
    ws2.cell(row=row, column=2, value=person).font = normal_font
    ws2.cell(row=row, column=3, value=status).font = normal_font
    for c in range(1, 4):
        ws2.cell(row=row, column=c).alignment = center_align
        ws2.cell(row=row, column=c).border = thin_border
    dv.add(ws2.cell(row=row, column=3))
    row += 1


# ========================================
# Sheet 3：图表
# ========================================
ws3 = wb.create_sheet("图表")
ws3.sheet_properties.tabColor = MINT

for i in range(1, 9):
    ws3.column_dimensions[get_column_letter(i)].width = 14

# 标题
ws3.merge_cells("A1:H1")
ws3["A1"].value = "图表展示"
ws3["A1"].font = title_font
ws3["A1"].alignment = center_align

ws3.merge_cells("A2:H2")
ws3["A2"].fill = PatternFill(start_color=TIFFANY, end_color=TIFFANY, fill_type="solid")

# 图表数据
chart_data = [
    ["月份", "销售额", "成本", "利润"],
    ["1月", 12000, 7200, 4800],
    ["2月", 15000, 8500, 6500],
    ["3月", 13500, 7800, 5700],
    ["4月", 18000, 9800, 8200],
    ["5月", 21000, 11000, 10000],
    ["6月", 19500, 10200, 9300],
]

row = 4
for i, h in enumerate(chart_data[0], 1):
    ws3.cell(row=row, column=i, value=h)
style_header_row(ws3, row, 4)
row += 1

for data_row in chart_data[1:]:
    for c_idx, val in enumerate(data_row):
        cell = ws3.cell(row=row, column=c_idx + 1, value=val)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
        if c_idx >= 1:
            cell.number_format = "#,##0"
    row += 1

data_end = row - 1

# --- 柱状图 ---
row += 1
write_section_title(ws3, row, "柱状图：销售额 vs 成本")
row += 1

bar_chart = BarChart()
bar_chart.type = "col"
bar_chart.title = "销售额 vs 成本"
bar_chart.y_axis.title = "金额"
bar_chart.x_axis.title = "月份"
bar_chart.style = 10
bar_chart.width = 18
bar_chart.height = 12

cats = Reference(ws3, min_col=1, min_row=5, max_row=data_end)
data1 = Reference(ws3, min_col=2, min_row=4, max_row=data_end)
data2 = Reference(ws3, min_col=3, min_row=4, max_row=data_end)

bar_chart.add_data(data1, titles_from_data=True)
bar_chart.add_data(data2, titles_from_data=True)
bar_chart.set_categories(cats)

bar_chart.series[0].graphicalProperties.solidFill = TIFFANY
bar_chart.series[1].graphicalProperties.solidFill = GRAY

ws3.add_chart(bar_chart, f"F4")

# --- 折线图 ---
row_chart_start = data_end + 18
write_section_title(ws3, row_chart_start, "折线图：利润趋势")
row_chart_start += 1

line_chart = LineChart()
line_chart.title = "利润趋势"
line_chart.y_axis.title = "利润"
line_chart.style = 10
line_chart.width = 18
line_chart.height = 12

data3 = Reference(ws3, min_col=4, min_row=4, max_row=data_end)
line_chart.add_data(data3, titles_from_data=True)
line_chart.set_categories(cats)

line_chart.series[0].graphicalProperties.line.solidFill = DEEP_TF
line_chart.series[0].graphicalProperties.line.width = 25000  # 2pt

ws3.add_chart(line_chart, f"F{row_chart_start}")

# --- 饼图 ---
pie_start = row_chart_start
write_section_title(ws3, pie_start, "饼图：各月销售额占比")
pie_start += 1

pie_chart = PieChart()
pie_chart.title = "各月销售额占比"
pie_chart.style = 10
pie_chart.width = 14
pie_chart.height = 12

data4 = Reference(ws3, min_col=2, min_row=4, max_row=data_end)
pie_chart.add_data(data4, titles_from_data=True)
pie_chart.set_categories(cats)

# 饼图颜色
colors_list = [TIFFANY, DEEP_TF, MINT, CREAM, GRAY, DARK]
for i, color in enumerate(colors_list):
    pt = DataPoint(idx=i)
    pt.graphicalProperties.solidFill = color
    pie_chart.series[0].data_points.append(pt)

pie_chart.dataLabels = DataLabelList()
pie_chart.dataLabels.showPercent = True
pie_chart.dataLabels.showVal = False

ws3.add_chart(pie_chart, f"F{pie_start}")


# ========================================
# Sheet 4：高级功能
# ========================================
ws4 = wb.create_sheet("高级功能")
ws4.sheet_properties.tabColor = DEEP_TF

for i in range(1, 9):
    ws4.column_dimensions[get_column_letter(i)].width = 16

# 标题
ws4.merge_cells("A1:H1")
ws4["A1"].value = "高级功能"
ws4["A1"].font = title_font
ws4["A1"].alignment = center_align

ws4.merge_cells("A2:H2")
ws4["A2"].fill = PatternFill(start_color=TIFFANY, end_color=TIFFANY, fill_type="solid")

# --- 冻结窗格 ---
row = 4
write_section_title(ws4, row, "冻结窗格（向下滚动时表头固定）")
row += 1

for i, h in enumerate(["姓名", "部门", "职位", "工资", "绩效", "奖金"], 1):
    ws4.cell(row=row, column=i, value=h)
style_header_row(ws4, row, 6)
ws4.freeze_panes = "A6"  # 冻结前5行
row += 1

staff = [
    ["张三", "技术部", "高级工程师", 25000, "A", 5000],
    ["李四", "产品部", "产品经理", 22000, "B+", 3000],
    ["王五", "设计部", "UI设计师", 18000, "A", 4000],
    ["赵六", "数据部", "数据分析师", 20000, "B", 2000],
    ["孙七", "技术部", "前端开发", 19000, "A-", 3500],
    ["周八", "运维部", "运维工程师", 17000, "B+", 2500],
    ["吴九", "市场部", "市场专员", 15000, "C", 1000],
    ["郑十", "财务部", "财务主管", 23000, "A", 6000],
    ["陈一", "技术部", "架构师", 30000, "S", 10000],
    ["林二", "产品部", "产品总监", 28000, "A+", 8000],
]

for s in staff:
    for c_idx, val in enumerate(s):
        cell = ws4.cell(row=row, column=c_idx + 1, value=val)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
        if c_idx == 3:
            cell.number_format = "¥#,##0"
        if c_idx == 5:
            cell.number_format = "¥#,##0"
    row += 1

row += 1

# --- 自动筛选 ---
write_section_title(ws4, row, "自动筛选（表头添加筛选箭头）")
row += 1
ws4.auto_filter.ref = f"A5:F{row - 3}"

# --- 打印设置 ---
row += 2
write_section_title(ws4, row, "打印设置")
row += 1

ws4.page_setup.orientation = "landscape"
ws4.page_setup.paperSize = ws4.PAPERSIZE_A4
ws4.page_margins.left = 0.6
ws4.page_margins.right = 0.6
ws4.page_margins.top = 0.8
ws4.page_margins.bottom = 0.8
ws4.print_title_rows = "1:5"  # 打印时重复前5行

ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
ws4.cell(row=row, column=1, value="横向打印 | A4纸 | 左右0.6英寸 | 重复前5行").font = gray_font
row += 1

# --- 公式 ---
row += 2
write_section_title(ws4, row, "常用公式示例")
row += 1

formulas = [
    ("求和", "=SUM(B2:B10)", "计算区域总和"),
    ("平均值", "=AVERAGE(B2:B10)", "计算平均值"),
    ("最大值", "=MAX(B2:B10)", "取最大值"),
    ("计数", "=COUNT(B2:B10)", "统计数字个数"),
    ("条件计数", '=COUNTIF(B2:B10,">20000")', "按条件统计"),
    ("查找", '=VLOOKUP("张三",A2:F10,4,FALSE)', "按行查找"),
    ("条件求和", '=SUMIF(C2:C10,"技术部",D2:D10)', "按条件求和"),
]

for i, h in enumerate(["公式名", "公式", "说明"], 1):
    ws4.cell(row=row, column=i, value=h)
style_header_row(ws4, row, 3)
row += 1

for name, formula, desc in formulas:
    ws4.cell(row=row, column=1, value=name).font = normal_font
    ws4.cell(row=row, column=2, value=formula).font = Font(name="Consolas", color=DEEP_TF, size=10)
    ws4.cell(row=row, column=3, value=desc).font = normal_font
    for c in range(1, 4):
        ws4.cell(row=row, column=c).alignment = center_align
        ws4.cell(row=row, column=c).border = thin_border
    row += 1


# ========================================
# Sheet 5：进度追踪
# ========================================
ws5 = wb.create_sheet("项目进度")
ws5.sheet_properties.tabColor = "E86D50"

for i in range(1, 10):
    ws5.column_dimensions[get_column_letter(i)].width = 14

# 标题
ws5.merge_cells("A1:I1")
ws5["A1"].value = "项目进度追踪"
ws5["A1"].font = title_font
ws5["A1"].alignment = center_align

ws5.merge_cells("A2:I2")
ws5["A2"].fill = PatternFill(start_color="E86D50", end_color="E86D50", fill_type="solid")

row = 4
headers = ["任务", "负责人", "开始日期", "截止日期", "进度", "状态", "优先级", "备注"]
for i, h in enumerate(headers, 1):
    ws5.cell(row=row, column=i, value=h)
for c in range(1, 9):
    cell = ws5.cell(row=row, column=c)
    cell.font = Font(name="Microsoft YaHei", bold=True, color=WHITE, size=11)
    cell.fill = PatternFill(start_color="E86D50", end_color="E86D50", fill_type="solid")
    cell.alignment = center_align
    cell.border = thin_border
row += 1

tasks = [
    ["需求分析", "张三", "2026-06-01", "2026-06-07", 100, "已完成", "P0", "已评审"],
    ["UI设计", "李四", "2026-06-05", "2026-06-14", 80, "进行中", "P1", "设计稿审核中"],
    ["数据库设计", "王五", "2026-06-08", "2026-06-12", 100, "已完成", "P0", ""],
    ["后端开发", "张三", "2026-06-10", "2026-06-25", 45, "进行中", "P0", "核心模块开发中"],
    ["前端开发", "赵六", "2026-06-12", "2026-06-28", 20, "进行中", "P1", ""],
    ["接口联调", "张三/赵六", "2026-06-22", "2026-06-30", 0, "未开始", "P0", ""],
    ["测试", "孙七", "2026-06-28", "2026-07-05", 0, "未开始", "P1", ""],
    ["上线部署", "周八", "2026-07-05", "2026-07-08", 0, "未开始", "P0", "灰度发布"],
]

for task_row in tasks:
    for c_idx, val in enumerate(task_row):
        cell = ws5.cell(row=row, column=c_idx + 1, value=val)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
        if c_idx == 4:  # 进度列
            cell.number_format = "0%"
            cell.value = val / 100
    row += 1

# 条件格式：状态列
status_end = row - 1
ws5.conditional_formatting.add(
    f"F5:F{status_end}",
    CellIsRule(operator="equal", formula=['"已完成"'],
               fill=PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
               font=Font(color="2E7D32", bold=True)),
)
ws5.conditional_formatting.add(
    f"F5:F{status_end}",
    CellIsRule(operator="equal", formula=['"进行中"'],
               fill=PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
               font=Font(color="E65100", bold=True)),
)
ws5.conditional_formatting.add(
    f"F5:F{status_end}",
    CellIsRule(operator="equal", formula=['"未开始"'],
               fill=PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
               font=Font(color="7B1FA2", bold=True)),
)

# 条件格式：进度列数据条
ws5.conditional_formatting.add(
    f"E5:E{status_end}",
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="81D8D1"),
)

# 优先级列着色
ws5.conditional_formatting.add(
    f"G5:G{status_end}",
    CellIsRule(operator="equal", formula=['"P0"'],
               font=Font(color="C62828", bold=True)),
)
ws5.conditional_formatting.add(
    f"G5:G{status_end}",
    CellIsRule(operator="equal", formula=['"P1"'],
               font=Font(color="E65100", bold=True)),
)

# 冻结窗格
ws5.freeze_panes = "A5"


# ========================================
# 保存
# ========================================
output_path = os.path.join(OUTPUT_DIR, "openpyxl_全功能展示.xlsx")
wb.save(output_path)
print(f"Excel 已生成: {output_path}")
print("openpyxl 全功能展示完成!")
